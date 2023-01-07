from tkinter import *
import tkinter as tk              
from tkinter import ttk
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

main=Tk()
main.geometry("800x500")
main.title("Giao Dịch")
tabControl = ttk.Notebook(main)
tabControl.pack()

tab1 = tk.Frame(tabControl,width = 800, height = 500)
tab2 = tk.Frame(tabControl)
tab1.pack(expand = 'yes', fill = 'both')
tab2.pack(expand = 'yes', fill = 'both')

tabControl.add(tab1, text ='Tab 1')
tabControl.add(tab2, text ='Tab 2')


# main.config(highlightbackground= "black", highlightthickness =2)

file = pathlib.Path("1.xlsx")
if file.exists ():
	pass
else:
	file = Workbook()
	sheet = file.active
	sheet["A1"]="Họ và tên"
	sheet["B1"]="Địa chỉ"
	sheet["C1"]="SĐT"
	sheet["D1"]="Mails"
	sheet["E1"]="abc"
	sheet["F1"]="xyz"

	file.save("1.xlsx")

def submit():
	y = name.get()
	z = user.get()
	z1 = passEntry.get()
	y1 = emailentry.get()

	file = openpyxl.load_workbook("1.xlsx")
	sheet = file.active
	sheet.cell(column = 1, row = sheet.max_row + 1, value = y)
	sheet.cell(column = 2, row = sheet.max_row, value = z)
	sheet.cell(column = 4, row = sheet.max_row, value = z1)
	sheet.cell(column = 3, row = sheet.max_row, value = y1)

	if var1.get() == 1:
		gen = "Male"
		sheet.cell(column = 5, row = sheet.max_row, value = "Nam")
	else:
		sheet.cell(column = 5, row = sheet.max_row, value = "Nữ")

	if var2.get() == 1:
		sheet.cell(column = 6, row = sheet.max_row, value = "Standard")

	if var3.get() == 1:
		sheet.cell(column = 6, row = sheet.max_row, value = "Premium")
	file.save("1.xlsx")



# frame1 = Label(tab1, text = 'Thông tin:').pack()

Label(tab1, text = "Họ và tên:").place(x = 50, y = 30)
Label(tab1, text = "Địa chỉ:").place(x = 50, y = 70)
Label(tab1, text = "SĐT").place(x = 50, y = 110)
Label(tab1, text = "Mail:").place(x = 50, y = 150)

Label(tab1, text = "Giới tính:").place(x = 50, y = 190)
Label(tab1, text = "Thứ hạng").place(x = 50, y = 230)

var1 = IntVar()
Radiobutton(tab1, text = "Nam", padx = 5, variable = var1, value = 1).place(x = 200, y = 190)
Radiobutton(tab1, text = "Nữ", padx = 20, variable = var1, value = 2).place(x = 280, y = 190)

frame2 = LabelFrame(tab2, text = 'Thông tin giao dịch:').pack(expand = 'yes', fill = 'both')

name = Entry(tab1)
name.place(x = 250, y = 30)
user = Entry(tab1)
user.place(x = 250, y = 70)
password = StringVar()
passEntry = Entry(tab1)
passEntry.place(x = 250 , y = 110)

emailentry = Entry(tab1)
emailentry.place(x = 250, y = 150, width = 250)

var = IntVar()
var.set('0')



var2 = IntVar()
Checkbutton(tab1, text = "Standard", variable = var2 ).place(x=200, y=230)
var3 = IntVar()
Checkbutton(tab1, text = "Premium", variable = var3 ).place(x=300, y=230)

Button(tab1, text = "submit", command = submit).place(x = 400, y = 400, anchor=SE)

main.mainloop()
